from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MISSING_FILL = PatternFill("solid", fgColor="FFFFF2CC")
REVIEW_FILL = PatternFill("solid", fgColor="FFFCE4D6")
HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
READ_ONLY_FILL = PatternFill("solid", fgColor="FFE7E6E6")
MAPPED_FILL = PatternFill("solid", fgColor="FFE2F0D9")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFD9E2F3"),
    right=Side(style="thin", color="FFD9E2F3"),
    top=Side(style="thin", color="FFD9E2F3"),
    bottom=Side(style="thin", color="FFD9E2F3"),
)


@dataclass(frozen=True)
class ConversionSummary:
    project_count: int
    quantity_rows: int
    collection_rows: int
    invoice_rows: int
    payment_rows: int
    unmatched_project_rows: int
    output_path: Path


@dataclass
class PreparedRow:
    values: list[object]
    marks: dict[int, tuple[str, str]]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\t", "").replace("\u00a0", " ").strip()


def normalize_project_name(value: object) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def parse_decimal(value: object) -> Decimal:
    text = clean_text(value).replace(",", "").replace("%", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def numeric_value(value: object) -> object:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return value
    text = clean_text(value).replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation:
        return value


def format_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    match = re.match(r"^(\d{4}-\d{2}-\d{2})", text)
    return match.group(1) if match else text


def simple_tax_rate(value: object) -> Decimal | None:
    match = re.fullmatch(r"\s*(\d+(?:\.\d+)?)%\s*", clean_text(value))
    return Decimal(match.group(1)) / Decimal("100") if match else None


def project_number(source_id: object) -> str:
    value = clean_text(source_id)
    return f"OLD-{value}" if value else ""


def source_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise ValueError(f"源工作簿没有表头：{path.name}")
    headers = [clean_text(value) for value in rows[0]]
    materialized: list[dict[str, object]] = []
    for row in rows[1:]:
        materialized.append({header: row[index] if index < len(row) else None for index, header in enumerate(headers)})
    return headers, materialized


def classify_workbooks(source_dir: Path, output_path: Path) -> dict[str, tuple[list[str], list[dict[str, object]]]]:
    required_signatures = {
        "projects": {"工程名称", "应收款", "甲方单位名称", "项目状态"},
        "quantities": {"工程名称", "工程量", "是否不开发票"},
        "collections": {"工程名称", "收款金额", "收款方式"},
        "invoices": {"工程名称", "开票金额", "开票日期"},
        "payments": {"工程名称", "小计", "扣款", "已付款", "未付款"},
    }
    found: dict[str, tuple[list[str], list[dict[str, object]]]] = {}
    for path in sorted(source_dir.glob("*.xlsx")):
        if path.resolve() == output_path.resolve():
            continue
        headers, rows = source_rows(path)
        header_set = set(headers)
        matches = [key for key, signature in required_signatures.items() if signature.issubset(header_set)]
        if len(matches) != 1:
            continue
        key = matches[0]
        if key in found:
            raise ValueError(f"发现重复的{key}源工作簿：{path.name}")
        found[key] = (headers, rows)
    missing = sorted(set(required_signatures) - set(found))
    if missing:
        raise ValueError(f"缺少源工作簿：{', '.join(missing)}")
    return found


def raw_columns(headers: Iterable[str]) -> list[str]:
    return [f"原始_{header}" for header in headers]


def raw_values(headers: Iterable[str], row: dict[str, object]) -> list[object]:
    return [row.get(header) for header in headers]


def add_mark(marks: dict[int, tuple[str, str]], headers: list[str], header: str, kind: str, reason: str) -> None:
    marks[headers.index(header)] = (kind, reason)


def append_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[PreparedRow], read_only_headers: set[str] | None = None) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(1, column)
        cell.fill = READ_ONLY_FILL if read_only_headers and header in read_only_headers else HEADER_FILL
        cell.font = Font(color="FFFFFFFF" if cell.fill == HEADER_FILL else "FF404040", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for prepared in rows:
        worksheet.append(prepared.values)
        row_number = worksheet.max_row
        for column in range(1, len(headers) + 1):
            cell = worksheet.cell(row_number, column)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
        for index, (kind, reason) in prepared.marks.items():
            cell = worksheet.cell(row_number, index + 1)
            cell.fill = MISSING_FILL if kind == "missing" else REVIEW_FILL
            cell.comment = Comment(reason, "Codex")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 34
    for index, header in enumerate(headers, start=1):
        sample = [clean_text(worksheet.cell(row, index).value) for row in range(1, min(worksheet.max_row, 80) + 1)]
        width = min(max([len(header), *(len(value) for value in sample)]) + 2, 42)
        if "备注" in header or "提示" in header or "说明" in header:
            width = min(max(width, 26), 52)
        worksheet.column_dimensions[get_column_letter(index)].width = width


def build_project_rows(
    headers: list[str], rows: list[dict[str, object]]
) -> tuple[list[PreparedRow], dict[str, str], dict[str, dict[str, object]]]:
    target_headers = [
        "项目编号", "项目名称", "项目阶段", "总包单位", "总包联系人", "税率", "税点原文",
        "项目状态原文", "签订合同公司原文", "项目经理原文", "施工机械原文", "工程量签认原文",
        "工程量附件文件名", "合同签定原文", "结算单签订原文", "结算单附件文件名",
        "付款条件原文", "备注", "来源数据唯一编号", "导入提示",
    ] + raw_columns(headers)
    stage_mapping = {
        "施工中": ("施工中", False),
        "阶段性施工暂停": ("停工中", True),
        "施工完成-未结清": ("已完工未结算", True),
        "施工完成-已结清": ("已结算归档", True),
    }
    prepared: list[PreparedRow] = []
    name_to_number: dict[str, str] = {}
    name_to_project: dict[str, dict[str, object]] = {}
    for row in rows:
        name = normalize_project_name(row.get("工程名称"))
        number = project_number(row.get("数据唯一编号"))
        if not name or not number:
            raise ValueError("项目统计表中的项目名称和数据唯一编号不能为空。")
        if name in name_to_number:
            raise ValueError(f"项目统计表中存在重复项目名称：{name}")
        name_to_number[name] = number
        name_to_project[name] = row
        source_stage = clean_text(row.get("项目状态"))
        stage, review_stage = stage_mapping.get(source_stage, ("", bool(source_stage)))
        tax_source = clean_text(row.get("税点"))
        tax_rate = simple_tax_rate(tax_source)
        values: list[object] = [
            number,
            name,
            stage,
            clean_text(row.get("甲方单位名称")),
            clean_text(row.get("甲方联系人")),
            tax_rate,
            tax_source,
            source_stage,
            clean_text(row.get("签订合同公司")),
            clean_text(row.get("项目经理")),
            clean_text(row.get("施工机械")),
            clean_text(row.get("工程量签认")),
            clean_text(row.get("工程量附件")),
            clean_text(row.get("合同签定")),
            clean_text(row.get("结算单签订")),
            clean_text(row.get("结算单附件")),
            clean_text(row.get("付款条件")),
            row.get("备注"),
            clean_text(row.get("数据唯一编号")),
            "",
            *raw_values(headers, row),
        ]
        marks: dict[int, tuple[str, str]] = {}
        prompts: list[str] = []
        if not source_stage:
            add_mark(marks, target_headers, "项目阶段", "missing", "源表未填写项目状态，请选择项目阶段。")
            prompts.append("补项目阶段")
        elif review_stage:
            add_mark(marks, target_headers, "项目阶段", "review", f"由“{source_stage}”映射为“{stage}”，请确认。")
            prompts.append("确认项目阶段")
        elif not stage:
            add_mark(marks, target_headers, "项目阶段", "review", f"项目状态“{source_stage}”无法自动映射。")
            prompts.append("确认项目阶段")
        if clean_text(row.get("签订合同公司")):
            add_mark(marks, target_headers, "签订合同公司原文", "review", "需要与系统自有公司编码核对。")
            prompts.append("核对签约公司")
        if clean_text(row.get("项目经理")):
            add_mark(marks, target_headers, "项目经理原文", "review", "需要与系统用户或员工主档核对。")
            prompts.append("核对项目经理")
        if tax_rate is None:
            kind = "missing" if not tax_source else "review"
            reason = "源表未填写税点，请补充。" if not tax_source else f"税点“{tax_source}”不是单一百分比，请确认税务配置。"
            add_mark(marks, target_headers, "税率", kind, reason)
            prompts.append("补充/确认税率")
        for target, source in (("工程量附件文件名", "工程量附件"), ("结算单附件文件名", "结算单附件")):
            if clean_text(row.get(source)):
                add_mark(marks, target_headers, target, "review", "源目录只有附件文件名，没有实际附件文件。")
                prompts.append("补附件")
        values[target_headers.index("导入提示")] = "；".join(dict.fromkeys(prompts))
        prepared.append(PreparedRow(values, marks))
    return prepared, name_to_number, name_to_project


def resolve_project(
    row: dict[str, object], name_to_number: dict[str, str]
) -> tuple[str, str, bool, str]:
    source_name = normalize_project_name(row.get("工程名称"))
    if not source_name:
        return "", "", False, "源表项目名称为空，请手动补充项目编号和项目名称。"
    number = name_to_number.get(source_name, "")
    if not number:
        return "", source_name, False, "项目名称未匹配项目统计表，请手动确认。"
    return number, source_name, True, ""


def build_quantity_rows(
    headers: list[str], rows: list[dict[str, object]], name_to_number: dict[str, str]
) -> tuple[list[PreparedRow], list[PreparedRow]]:
    target_headers = [
        "项目编号", "项目名称", "工程量名称", "工程量", "单位", "单价", "金额", "金额校验差额",
        "是否不开发票", "附件文件名", "备注", "来源数据唯一编号", "导入目标（待确认）", "导入提示",
    ] + raw_columns(headers)
    prepared: list[PreparedRow] = []
    exceptions: list[PreparedRow] = []
    for row_number, row in enumerate(rows, start=2):
        number, name, matched, project_issue = resolve_project(row, name_to_number)
        quantity = parse_decimal(row.get("工程量"))
        unit_price = parse_decimal(row.get("单价"))
        amount = parse_decimal(row.get("小计"))
        difference = amount - quantity * unit_price
        values: list[object] = [
            number, name, clean_text(row.get("名称")), numeric_value(row.get("工程量")), clean_text(row.get("单位")),
            numeric_value(row.get("单价")), numeric_value(row.get("小计")), difference,
            clean_text(row.get("是否不开发票")), clean_text(row.get("附件")), row.get("备注"),
            clean_text(row.get("数据唯一编号")), "", "", *raw_values(headers, row),
        ]
        marks: dict[int, tuple[str, str]] = {}
        prompts = ["确认导入目标"]
        add_mark(marks, target_headers, "导入目标（待确认）", "review", "当前系统没有独立工程量数据集，请确认导入为合同清单项或其他业务明细。")
        if not matched:
            add_mark(marks, target_headers, "项目编号", "review", project_issue)
            add_mark(marks, target_headers, "项目名称", "review", project_issue)
            prompts.append("补项目")
            exceptions.append(exception_row("工程量导入", row_number, row, amount, project_issue))
        if abs(difference) > Decimal("0.01"):
            add_mark(marks, target_headers, "金额校验差额", "review", "小计不等于工程量乘单价，请确认原始金额。")
            prompts.append("核对金额")
        if clean_text(row.get("附件")):
            add_mark(marks, target_headers, "附件文件名", "review", "源目录只有附件文件名，没有实际附件文件。")
            prompts.append("补附件")
        values[target_headers.index("导入提示")] = "；".join(dict.fromkeys(prompts))
        prepared.append(PreparedRow(values, marks))
    return prepared, exceptions


def mapped_payment_method(source: object) -> tuple[str, bool]:
    value = clean_text(source)
    mapping = {
        "对公现金汇入": "BankTransfer",
        "对私汇入": "BankTransfer",
        "对公承兑汇入": "Other",
        "民工工资": "Other",
        "其他": "Other",
    }
    return mapping.get(value, ""), bool(value)


def build_collection_rows(
    headers: list[str], rows: list[dict[str, object]], name_to_number: dict[str, str]
) -> tuple[list[PreparedRow], list[PreparedRow]]:
    target_headers = [
        "项目编号", "项目名称", "收款日期", "收款金额", "收款方式（系统值）", "收款方式原文", "是否开票原文",
        "公司编码（待补）", "账户编号（待补）", "往来单位编号（待补）", "附件文件名", "备注",
        "来源数据唯一编号", "导入提示",
    ] + raw_columns(headers)
    prepared: list[PreparedRow] = []
    exceptions: list[PreparedRow] = []
    for row_number, row in enumerate(rows, start=2):
        number, name, matched, project_issue = resolve_project(row, name_to_number)
        method, inferred = mapped_payment_method(row.get("收款方式"))
        values: list[object] = [
            number, name, format_date(row.get("收款日期")), numeric_value(row.get("收款金额")), method,
            clean_text(row.get("收款方式")), clean_text(row.get("是否开票")), "", "", "",
            clean_text(row.get("附件")), row.get("备注"), clean_text(row.get("数据唯一编号")), "", *raw_values(headers, row),
        ]
        marks: dict[int, tuple[str, str]] = {}
        prompts = ["补公司编码", "补账户编号", "补往来单位"]
        for target, reason in (
            ("公司编码（待补）", "收款记录必须选择收款公司。"),
            ("账户编号（待补）", "收款记录必须选择收款账户。"),
            ("往来单位编号（待补）", "源表没有往来单位编号，请补充或确认留空。"),
        ):
            add_mark(marks, target_headers, target, "missing", reason)
        if not matched:
            add_mark(marks, target_headers, "项目编号", "review", project_issue)
            add_mark(marks, target_headers, "项目名称", "review", project_issue)
            prompts.append("补项目")
            exceptions.append(exception_row("收款导入", row_number, row, parse_decimal(row.get("收款金额")), project_issue))
        if inferred:
            add_mark(marks, target_headers, "收款方式（系统值）", "review", f"由原值“{clean_text(row.get('收款方式'))}”推断，请确认。")
            prompts.append("确认收款方式")
        else:
            add_mark(marks, target_headers, "收款方式（系统值）", "missing", "源表未填写收款方式，请补充。")
            prompts.append("补收款方式")
        if not format_date(row.get("收款日期")):
            add_mark(marks, target_headers, "收款日期", "missing", "收款日期不能为空。")
            prompts.append("补收款日期")
        if clean_text(row.get("附件")):
            add_mark(marks, target_headers, "附件文件名", "review", "源目录只有附件文件名，没有实际附件文件。")
            prompts.append("补附件")
        values[target_headers.index("导入提示")] = "；".join(dict.fromkeys(prompts))
        prepared.append(PreparedRow(values, marks))
    return prepared, exceptions


def build_invoice_rows(
    headers: list[str], rows: list[dict[str, object]], name_to_number: dict[str, str], name_to_project: dict[str, dict[str, object]]
) -> tuple[list[PreparedRow], list[PreparedRow]]:
    target_headers = [
        "项目编号", "项目名称", "发票号码（待补）", "开票日期", "发票方向（待确认）", "公司编码（待补）",
        "往来单位编号（待补）", "税率（待确认）", "税点原文", "不含税金额（待确认）", "税额（待确认）",
        "价税合计", "发票状态（待确认）", "附件文件名", "备注", "来源数据唯一编号", "导入提示",
    ] + raw_columns(headers)
    prepared: list[PreparedRow] = []
    exceptions: list[PreparedRow] = []
    for row_number, row in enumerate(rows, start=2):
        number, name, matched, project_issue = resolve_project(row, name_to_number)
        gross = parse_decimal(row.get("开票金额"))
        project = name_to_project.get(name)
        tax_source = clean_text(project.get("税点")) if project else ""
        tax_rate = simple_tax_rate(tax_source)
        net = (gross / (Decimal("1") + tax_rate)).quantize(Decimal("0.01")) if tax_rate is not None else None
        tax = gross - net if net is not None else None
        values: list[object] = [
            number, name, "", format_date(row.get("开票日期")), "Output", "", "", tax_rate, tax_source,
            net, tax, numeric_value(row.get("开票金额")), "IssuedOrReceived", clean_text(row.get("附件")), row.get("备注"),
            clean_text(row.get("数据唯一编号")), "", *raw_values(headers, row),
        ]
        marks: dict[int, tuple[str, str]] = {}
        prompts = ["补发票号码", "确认发票方向", "补公司编码", "补往来单位", "确认发票状态"]
        add_mark(marks, target_headers, "发票号码（待补）", "missing", "源表没有发票号码，请补充。")
        add_mark(marks, target_headers, "发票方向（待确认）", "review", "按项目开票表推断为销项发票，请确认。")
        add_mark(marks, target_headers, "公司编码（待补）", "missing", "发票必须选择开票公司。")
        add_mark(marks, target_headers, "往来单位编号（待补）", "missing", "源表没有往来单位编号，请补充。")
        add_mark(marks, target_headers, "发票状态（待确认）", "review", "按历史开票记录推断为已开具/已收到，请确认。")
        if tax_rate is None:
            kind = "missing" if not tax_source else "review"
            reason = "项目未填写单一税率，请补充发票税率。" if not tax_source else f"项目税点“{tax_source}”不能自动确定单一税率。"
            for target in ("税率（待确认）", "不含税金额（待确认）", "税额（待确认）"):
                add_mark(marks, target_headers, target, kind, reason)
            prompts.append("补税率和税额")
        else:
            for target in ("税率（待确认）", "不含税金额（待确认）", "税额（待确认）"):
                add_mark(marks, target_headers, target, "review", "根据项目税点和价税合计推算，请确认。")
            prompts.append("确认税率和税额")
        if not matched:
            add_mark(marks, target_headers, "项目编号", "review", project_issue)
            add_mark(marks, target_headers, "项目名称", "review", project_issue)
            prompts.append("补项目")
            exceptions.append(exception_row("开票导入", row_number, row, gross, project_issue))
        if not format_date(row.get("开票日期")):
            add_mark(marks, target_headers, "开票日期", "missing", "开票日期不能为空。")
            prompts.append("补开票日期")
        if clean_text(row.get("附件")):
            add_mark(marks, target_headers, "附件文件名", "review", "源目录只有附件文件名，没有实际附件文件。")
            prompts.append("补附件")
        values[target_headers.index("导入提示")] = "；".join(dict.fromkeys(prompts))
        prepared.append(PreparedRow(values, marks))
    return prepared, exceptions


def build_payment_rows(
    headers: list[str], rows: list[dict[str, object]], name_to_number: dict[str, str]
) -> tuple[list[PreparedRow], list[PreparedRow]]:
    target_headers = [
        "项目编号", "项目名称", "应付金额", "扣款金额", "已付款金额（原始）", "未付款金额", "付款日期（待补）",
        "付款金额（待确认）", "付款方式（待补）", "公司编码（待补）", "账户编号（待补）",
        "往来单位编号（待补）", "往来单位名称（原文）", "收款人", "已开票原文", "附件文件名", "备注",
        "来源数据唯一编号", "金额校验差额", "导入提示",
    ] + raw_columns(headers)
    prepared: list[PreparedRow] = []
    exceptions: list[PreparedRow] = []
    for row_number, row in enumerate(rows, start=2):
        number, name, matched, project_issue = resolve_project(row, name_to_number)
        subtotal = parse_decimal(row.get("小计"))
        deduction = parse_decimal(row.get("扣款"))
        paid = parse_decimal(row.get("已付款"))
        unpaid = parse_decimal(row.get("未付款"))
        difference = subtotal - deduction - paid - unpaid
        values: list[object] = [
            number, name, numeric_value(row.get("小计")), numeric_value(row.get("扣款")), numeric_value(row.get("已付款")),
            numeric_value(row.get("未付款")), "", paid if paid else None, "", "", "", "",
            clean_text(row.get("名称")), clean_text(row.get("收款人")), row.get("已开票"), clean_text(row.get("附件")),
            row.get("备注"), clean_text(row.get("数据唯一编号")), difference, "", *raw_values(headers, row),
        ]
        marks: dict[int, tuple[str, str]] = {}
        prompts = ["补公司编码", "补往来单位"]
        add_mark(marks, target_headers, "公司编码（待补）", "missing", "应付记录必须选择付款公司。")
        add_mark(marks, target_headers, "往来单位编号（待补）", "missing", "源表只有往来单位名称，请补充系统单位编号。")
        if paid:
            prompts.extend(["补付款日期", "确认付款金额", "补付款方式", "补账户编号"])
            add_mark(marks, target_headers, "付款日期（待补）", "missing", "源表只有累计已付款金额，没有付款日期。")
            add_mark(marks, target_headers, "付款金额（待确认）", "review", "暂按累计已付款金额填入；请确认是否应拆成多笔付款。")
            add_mark(marks, target_headers, "付款方式（待补）", "missing", "源表没有付款方式。")
            add_mark(marks, target_headers, "账户编号（待补）", "missing", "付款记录必须选择付款账户。")
        if not matched:
            add_mark(marks, target_headers, "项目编号", "review", project_issue)
            add_mark(marks, target_headers, "项目名称", "review", project_issue)
            prompts.append("补项目")
            exceptions.append(exception_row("付款导入", row_number, row, subtotal, project_issue))
        if abs(difference) > Decimal("0.01"):
            add_mark(marks, target_headers, "金额校验差额", "review", "小计减扣款、已付款和未付款不为零，请核对原始金额。")
            prompts.append("核对金额")
        if clean_text(row.get("已开票")) and parse_decimal(row.get("已开票")) == 0:
            add_mark(marks, target_headers, "已开票原文", "review", "已开票列不是有效金额，请确认源数据是否错列。")
            prompts.append("核对已开票")
        if clean_text(row.get("附件")):
            add_mark(marks, target_headers, "附件文件名", "review", "源目录只有附件文件名，没有实际附件文件。")
            prompts.append("补附件")
        values[target_headers.index("导入提示")] = "；".join(dict.fromkeys(prompts))
        prepared.append(PreparedRow(values, marks))
    return prepared, exceptions


def exception_row(source_sheet: str, row_number: int, row: dict[str, object], amount: Decimal, issue: str) -> PreparedRow:
    headers = ["来源工作表", "来源行号", "来源数据唯一编号", "原始项目名称", "项目编号（待补）", "项目名称（待补）", "金额", "名称/备注", "处理说明"]
    values: list[object] = [
        source_sheet, row_number, clean_text(row.get("数据唯一编号")), clean_text(row.get("工程名称")), "", "", amount,
        clean_text(row.get("名称")) or clean_text(row.get("备注")), issue,
    ]
    marks = {
        headers.index("项目编号（待补）"): ("review", issue),
        headers.index("项目名称（待补）"): ("review", issue),
    }
    return PreparedRow(values, marks)


def build_balance_rows(
    project_rows: list[dict[str, object]], name_to_number: dict[str, str],
    collection_rows: list[dict[str, object]], invoice_rows: list[dict[str, object]], payment_rows: list[dict[str, object]],
) -> list[PreparedRow]:
    collections: dict[str, Decimal] = defaultdict(Decimal)
    invoices: dict[str, Decimal] = defaultdict(Decimal)
    payments: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    for row in collection_rows:
        name = normalize_project_name(row.get("工程名称"))
        if name in name_to_number:
            collections[name] += parse_decimal(row.get("收款金额"))
    for row in invoice_rows:
        name = normalize_project_name(row.get("工程名称"))
        if name in name_to_number:
            invoices[name] += parse_decimal(row.get("开票金额"))
    for row in payment_rows:
        name = normalize_project_name(row.get("工程名称"))
        if name in name_to_number:
            payments[name]["subtotal"] += parse_decimal(row.get("小计"))
            payments[name]["deduction"] += parse_decimal(row.get("扣款"))
            payments[name]["paid"] += parse_decimal(row.get("已付款"))
            payments[name]["unpaid"] += parse_decimal(row.get("未付款"))
            payments[name]["invoiced"] += parse_decimal(row.get("已开票"))
    headers = [
        "项目编号", "项目名称", "统计应收款", "统计已收款", "统计未收款", "统计付款比例", "统计已开票", "统计未开票",
        "收款明细合计", "已收款对账差额", "开票明细合计", "已开票对账差额", "付款明细小计", "付款明细扣款",
        "付款明细已付款", "付款明细未付款", "付款明细已开票", "来源数据唯一编号", "备注",
    ]
    result: list[PreparedRow] = []
    for row in project_rows:
        name = normalize_project_name(row.get("工程名称"))
        collected_snapshot = parse_decimal(row.get("已收款"))
        invoiced_snapshot = parse_decimal(row.get("已开票"))
        collection_difference = collected_snapshot - collections[name]
        invoice_difference = invoiced_snapshot - invoices[name]
        values: list[object] = [
            name_to_number[name], name, numeric_value(row.get("应收款")), numeric_value(row.get("已收款")),
            numeric_value(row.get("未收款")), numeric_value(row.get("付款比例")), numeric_value(row.get("已开票")),
            numeric_value(row.get("未开票")), collections[name], collection_difference, invoices[name], invoice_difference,
            payments[name]["subtotal"], payments[name]["deduction"], payments[name]["paid"], payments[name]["unpaid"],
            payments[name]["invoiced"], clean_text(row.get("数据唯一编号")), row.get("备注"),
        ]
        marks: dict[int, tuple[str, str]] = {}
        if abs(collection_difference) > Decimal("0.01"):
            add_mark(marks, headers, "已收款对账差额", "review", "项目统计表已收款与收款明细合计不一致，请确认迁移截止口径。")
        if abs(invoice_difference) > Decimal("0.01"):
            add_mark(marks, headers, "已开票对账差额", "review", "项目统计表已开票与开票明细合计不一致，请确认迁移截止口径。")
        result.append(PreparedRow(values, marks))
    return result


def build_intro_sheet(workbook: Workbook, summary_values: dict[str, int]) -> None:
    worksheet = workbook.active
    worksheet.title = "导入说明"
    worksheet.append(["项目", "说明"])
    rows = [
        ("转换范围", "已排除用户删除的班组付款表；本工作簿来自项目统计、工程量、收款、开票、付款共 5 个源文件。"),
        ("项目合并", f"以项目统计表 {summary_values['projects']} 个唯一项目名称为基准；项目编号为 OLD-加原数据唯一编号。"),
        ("明细行数", f"工程量 {summary_values['quantities']} 行，收款 {summary_values['collections']} 行，开票 {summary_values['invoices']} 行，付款 {summary_values['payments']} 行。"),
        ("待补项目", f"共有 {summary_values['unmatched']} 行缺少或未匹配项目名称，已集中到“待补项目”工作表，同时保留在原导入表。"),
        ("黄色单元格", "目标系统必填或业务必要字段在源表中缺失，需要手工补充。"),
        ("橙色单元格", "自动映射不完全等价、金额对账不一致、项目未匹配或附件文件缺失，需要人工确认。"),
        ("绿色说明", "字段已按明确规则自动映射。"),
        ("导入顺序", "项目导入 -> 工程量导入（确认目标后）-> 收款导入 -> 开票导入 -> 付款导入。导入前应处理所有黄色和橙色标记。"),
        ("余额快照", "项目余额快照仅用于迁移前后对账，不应与收款、开票、付款明细重复导入。"),
        ("附件", "当前 old-data 目录没有实际 PDF/JPG 附件；工作簿只保留附件文件名，相关单元格已标橙色。"),
        ("原始字段", "每个导入工作表末尾均保留“原始_”字段，确保源数据不因转换而丢失。"),
    ]
    for row in rows:
        worksheet.append(row)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    for row in range(2, worksheet.max_row + 1):
        for column in range(1, 3):
            worksheet.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)
            worksheet.cell(row, column).border = THIN_BORDER
    for label, fill in (("黄色单元格", MISSING_FILL), ("橙色单元格", REVIEW_FILL), ("绿色说明", MAPPED_FILL)):
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row, 1).value == label:
                worksheet.cell(row, 1).fill = fill
                worksheet.cell(row, 2).fill = fill
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 90
    worksheet.freeze_panes = "A2"


def convert_directory(source_dir: Path, output_path: Path) -> ConversionSummary:
    source_dir = source_dir.resolve()
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    books = classify_workbooks(source_dir, output_path)
    project_headers, project_source = books["projects"]
    quantity_headers, quantity_source = books["quantities"]
    collection_headers, collection_source = books["collections"]
    invoice_headers, invoice_source = books["invoices"]
    payment_headers, payment_source = books["payments"]

    projects, name_to_number, name_to_project = build_project_rows(project_headers, project_source)
    quantities, quantity_exceptions = build_quantity_rows(quantity_headers, quantity_source, name_to_number)
    collections, collection_exceptions = build_collection_rows(collection_headers, collection_source, name_to_number)
    invoices, invoice_exceptions = build_invoice_rows(invoice_headers, invoice_source, name_to_number, name_to_project)
    payments, payment_exceptions = build_payment_rows(payment_headers, payment_source, name_to_number)
    exceptions = quantity_exceptions + collection_exceptions + invoice_exceptions + payment_exceptions
    balances = build_balance_rows(project_source, name_to_number, collection_source, invoice_source, payment_source)

    workbook = Workbook()
    build_intro_sheet(workbook, {
        "projects": len(projects),
        "quantities": len(quantities),
        "collections": len(collections),
        "invoices": len(invoices),
        "payments": len(payments),
        "unmatched": len(exceptions),
    })
    append_sheet(workbook, "项目导入", [
        "项目编号", "项目名称", "项目阶段", "总包单位", "总包联系人", "税率", "税点原文",
        "项目状态原文", "签订合同公司原文", "项目经理原文", "施工机械原文", "工程量签认原文",
        "工程量附件文件名", "合同签定原文", "结算单签订原文", "结算单附件文件名",
        "付款条件原文", "备注", "来源数据唯一编号", "导入提示", *raw_columns(project_headers),
    ], projects)
    balance_headers = [
        "项目编号", "项目名称", "统计应收款", "统计已收款", "统计未收款", "统计付款比例", "统计已开票", "统计未开票",
        "收款明细合计", "已收款对账差额", "开票明细合计", "已开票对账差额", "付款明细小计", "付款明细扣款",
        "付款明细已付款", "付款明细未付款", "付款明细已开票", "来源数据唯一编号", "备注",
    ]
    append_sheet(workbook, "项目余额快照", balance_headers, balances, set(balance_headers))
    append_sheet(workbook, "工程量导入", [
        "项目编号", "项目名称", "工程量名称", "工程量", "单位", "单价", "金额", "金额校验差额",
        "是否不开发票", "附件文件名", "备注", "来源数据唯一编号", "导入目标（待确认）", "导入提示",
        *raw_columns(quantity_headers),
    ], quantities)
    append_sheet(workbook, "收款导入", [
        "项目编号", "项目名称", "收款日期", "收款金额", "收款方式（系统值）", "收款方式原文", "是否开票原文",
        "公司编码（待补）", "账户编号（待补）", "往来单位编号（待补）", "附件文件名", "备注",
        "来源数据唯一编号", "导入提示", *raw_columns(collection_headers),
    ], collections)
    append_sheet(workbook, "开票导入", [
        "项目编号", "项目名称", "发票号码（待补）", "开票日期", "发票方向（待确认）", "公司编码（待补）",
        "往来单位编号（待补）", "税率（待确认）", "税点原文", "不含税金额（待确认）", "税额（待确认）",
        "价税合计", "发票状态（待确认）", "附件文件名", "备注", "来源数据唯一编号", "导入提示",
        *raw_columns(invoice_headers),
    ], invoices)
    append_sheet(workbook, "付款导入", [
        "项目编号", "项目名称", "应付金额", "扣款金额", "已付款金额（原始）", "未付款金额", "付款日期（待补）",
        "付款金额（待确认）", "付款方式（待补）", "公司编码（待补）", "账户编号（待补）",
        "往来单位编号（待补）", "往来单位名称（原文）", "收款人", "已开票原文", "附件文件名", "备注",
        "来源数据唯一编号", "金额校验差额", "导入提示", *raw_columns(payment_headers),
    ], payments)
    append_sheet(workbook, "待补项目", [
        "来源工作表", "来源行号", "来源数据唯一编号", "原始项目名称", "项目编号（待补）", "项目名称（待补）",
        "金额", "名称/备注", "处理说明",
    ], exceptions)

    workbook.save(output_path)
    return ConversionSummary(
        project_count=len(projects),
        quantity_rows=len(quantities),
        collection_rows=len(collections),
        invoice_rows=len(invoices),
        payment_rows=len(payments),
        unmatched_project_rows=len(exceptions),
        output_path=output_path,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert legacy project Excel exports into a reviewed import workbook.")
    parser.add_argument("source_dir", type=Path)
    parser.add_argument("output_path", type=Path)
    args = parser.parse_args()
    summary = convert_directory(args.source_dir, args.output_path)
    print(json.dumps({
        "project_count": summary.project_count,
        "quantity_rows": summary.quantity_rows,
        "collection_rows": summary.collection_rows,
        "invoice_rows": summary.invoice_rows,
        "payment_rows": summary.payment_rows,
        "unmatched_project_rows": summary.unmatched_project_rows,
        "output_path": str(summary.output_path),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
