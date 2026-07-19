import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


SCRIPTS_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SCRIPTS_DIR))

from convert_legacy_project_data import convert_directory


def write_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Worksheet"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


class LegacyProjectDataConversionTests(unittest.TestCase):
    def test_conversion_preserves_rows_and_marks_missing_and_unmatched_cells(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory)
            output = source / "converted.xlsx"
            write_workbook(
                source / "project-summary.xlsx",
                ["数据唯一编号", "工程名称", "应收款", "已收款", "未收款", "付款比例", "已开票", "未开票", "甲方联系人", "甲方单位名称", "项目经理", "签订合同公司", "税点", "施工机械", "工程量签认", "工程量附件", "合同签定", "结算单签订", "结算单附件", "项目状态", "付款条件", "备注"],
                [
                    ["1001\t", "项目甲", 100, 20, 80, 20, 10, 90, "联系人", "甲方单位", "经理", "联蒲", "3%", None, None, None, "合同已签完", "未结算", None, "施工中", None, "备注"],
                    ["1002\t", "项目乙", 200, 200, 0, 100, 200, 0, None, None, None, None, None, None, None, None, None, None, None, "施工完成-已结清", None, None],
                ],
            )
            write_workbook(
                source / "quantity.xlsx",
                ["数据唯一编号", "工程名称", "名称", "工程量", "单位", "单价", "小计", "附件", "备注", "是否不开发票"],
                [["q1\t", "项目甲", "工程量一", 2, "项", 50, 100, None, None, None], ["q2\t", None, "待补项目", 1, "项", 20, 20, None, None, None]],
            )
            write_workbook(
                source / "invoice.xlsx",
                ["数据唯一编号", "工程名称", "开票金额", "开票日期", "附件", "备注"],
                [["i1\t", "项目甲", 10, "2026-01-01 00:00:00", None, None]],
            )
            write_workbook(
                source / "collection.xlsx",
                ["数据唯一编号", "工程名称", "收款金额", "收款日期", "收款方式", "是否开票", "附件", "备注"],
                [["c1\t", "项目甲", 20, "2026-01-02 00:00:00", "对公现金汇入", "开票", None, None]],
            )
            write_workbook(
                source / "payment.xlsx",
                ["数据唯一编号", "工程名称", "班组名称", "名称", "数量", "单位", "单价", "小计", "扣款", "已付款", "未付款", "已开票", "收款人", "附件", "备注"],
                [["p1\t", None, None, "付款项", 1, "项", 30, 30, None, None, 30, None, "收款人", None, None]],
            )

            summary = convert_directory(source, output)

            self.assertEqual(2, summary.project_count)
            self.assertEqual(2, summary.unmatched_project_rows)
            self.assertTrue(output.exists())

            workbook = load_workbook(output)
            self.assertEqual(
                ["导入说明", "项目导入", "项目余额快照", "工程量导入", "收款导入", "开票导入", "付款导入", "待补项目"],
                workbook.sheetnames,
            )
            self.assertEqual("OLD-1001", workbook["项目导入"]["A2"].value)

            collection = workbook["收款导入"]
            collection_headers = {cell.value: cell.column for cell in collection[1]}
            company_cell = collection.cell(2, collection_headers["公司编码（待补）"])
            self.assertEqual("FFFFF2CC", company_cell.fill.fgColor.rgb)

            quantity = workbook["工程量导入"]
            quantity_headers = {cell.value: cell.column for cell in quantity[1]}
            missing_project_cell = quantity.cell(3, quantity_headers["项目编号"])
            self.assertEqual("FFFCE4D6", missing_project_cell.fill.fgColor.rgb)
            self.assertEqual(2, quantity.max_row - 1)
            self.assertEqual(1, workbook["收款导入"].max_row - 1)
            self.assertEqual(1, workbook["开票导入"].max_row - 1)
            self.assertEqual(1, workbook["付款导入"].max_row - 1)


if __name__ == "__main__":
    unittest.main()
